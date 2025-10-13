import csv
import io
from datetime import datetime, timezone
from typing import Optional

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import and_
from sqlalchemy.orm import Session
from src.models.activity import Activity, Collaborator
from src.models.debate import Debate
from src.models.vote import Participant, Vote
from src.schemas.activity import CollaboratorStatus
from src.schemas.participant import (PaginatedParticipants,
                                     ParticipantBatchImportResult,
                                     ParticipantCreate, ParticipantResponse)


class ParticipantService:
    def __init__(self, db: Session):
        self.db = db

    def _check_activity_permission(self, activity_id: str, user_id: str) -> Activity:
        """检查用户对活动的权限"""
        activity = self.db.query(Activity).filter(
            Activity.id == activity_id).first()
        if not activity:
            raise HTTPException(status_code=404, detail="Activity not found")

        # 检查是否是活动拥有者或协作者
        if str(activity.owner_id) != str(user_id):
            # 检查是否是已接受的协作者
            collaborator = self.db.query(Collaborator).filter(
                Collaborator.activity_id == activity_id,
                Collaborator.user_id == user_id,
                Collaborator.status == CollaboratorStatus.accepted
            ).first()

            if not collaborator:
                raise HTTPException(
                    status_code=403,
                    detail="You don't have permission to access this activity"
                )

        return activity

    def get_participants_paginated(
        self,
        activity_id: str,
        user_id: str,
        page: int = 1,
        limit: int = 50,
        status: Optional[str] = None,
        search: Optional[str] = None,
        name: Optional[str] = None,
        code: Optional[str] = None,
        phone: Optional[str] = None,
        note: Optional[str] = None,
        checked_in: Optional[bool] = None,
        sort_by: Optional[str] = "created_at",
        sort_order: Optional[str] = "desc"
    ) -> PaginatedParticipants:
        """获取分页参与者列表"""
        # 检查权限
        self._check_activity_permission(activity_id, user_id)

        # 构建查询
        query = self.db.query(Participant).filter(
            Participant.activity_id == activity_id)

        # 状态筛选（向后兼容）
        if status == "checked_in":
            query = query.filter(Participant.checked_in == True)
        elif status == "not_checked_in":
            query = query.filter(Participant.checked_in == False)

        # 入场状态筛选（新方式）
        if checked_in is not None:
            query = query.filter(Participant.checked_in == checked_in)

        # 全文搜索
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                Participant.name.ilike(search_term) |
                Participant.code.ilike(search_term) |
                Participant.phone.ilike(search_term)
            )

        # 具体字段模糊匹配
        if name:
            query = query.filter(Participant.name.ilike(f"%{name}%"))

        if code:
            query = query.filter(Participant.code.ilike(f"%{code}%"))

        if phone:
            query = query.filter(Participant.phone.ilike(f"%{phone}%"))

        if note:
            query = query.filter(Participant.note.ilike(f"%{note}%"))

        # 排序
        if sort_by and hasattr(Participant, sort_by):
            sort_column = getattr(Participant, sort_by)
        else:
            sort_column = Participant.created_at

        if sort_order == "desc":
            query = query.order_by(sort_column.desc())
        else:
            query = query.order_by(sort_column.asc())

        # 计算总数
        total = query.count()

        # 分页
        offset = (page - 1) * limit
        participants = query.offset(offset).limit(limit).all()

        # 计算总页数
        total_pages = (total + limit - 1) // limit

        return PaginatedParticipants(
            items=[ParticipantResponse.model_validate(
                p) for p in participants],
            total=total,
            page=page,
            limit=limit,
            totalPages=total_pages
        )

    def create_participant(
        self,
        activity_id: str,
        participant_data: ParticipantCreate,
        user_id: str
    ) -> ParticipantResponse:
        """创建参与者"""
        # 检查权限
        self._check_activity_permission(activity_id, user_id)

        # 生成参与者编号
        code = self._generate_participant_code(activity_id)

        # 创建参与者
        participant = Participant(
            activity_id=activity_id,
            code=code,
            name=participant_data.name,
            phone=participant_data.phone,
            note=participant_data.note
        )

        self.db.add(participant)
        self.db.commit()
        self.db.refresh(participant)

        return ParticipantResponse.model_validate(participant)

    def _generate_participant_code(self, activity_id: str) -> str:
        """生成参与者编号"""
        # 获取当前活动的参与者数量
        count = self.db.query(Participant).filter(
            Participant.activity_id == activity_id
        ).count()
        return f"{count + 1:04d}"  # 生成4位数字编号，如0001, 0002

    def _detect_csv_format(self, header_row: list) -> dict:
        """智能识别CSV文件格式，返回列索引映射

        支持的格式：
        1. 导入模板格式：姓名,手机号,备注
        2. 导出文件格式：编号,姓名,手机号,备注,是否入场,入场时间,创建时间

        Returns:
            dict: {'id': int, 'name': int, 'phone': int, 'note': int} 列索引映射
        """
        # 标准化标题（去除空格和引号）
        headers = [h.strip().strip('"').strip("'") for h in header_row]

        # 初始化列映射
        column_mapping = {
            'id': -1,      # 编号/ID
            'name': -1,    # 姓名
            'phone': -1,   # 手机号
            'note': -1     # 备注
        }

        # 可能的列名匹配
        id_keywords = ['编号', 'id', 'code', '序号', '参与者编号']
        name_keywords = ['姓名', 'name', '名字', '参与者']
        phone_keywords = ['手机', 'phone', '电话', '手机号', '联系方式', '联系电话']
        note_keywords = ['备注', 'note', '说明', '备注信息', '描述']

        # 查找每个字段的列索引
        for idx, header in enumerate(headers):
            header_lower = header.lower()

            # 匹配编号列
            if column_mapping['id'] == -1:
                for keyword in id_keywords:
                    if keyword in header or keyword in header_lower:
                        column_mapping['id'] = idx
                        break

            # 匹配姓名列
            if column_mapping['name'] == -1:
                for keyword in name_keywords:
                    if keyword in header or keyword in header_lower:
                        column_mapping['name'] = idx
                        break

            # 匹配手机号列
            if column_mapping['phone'] == -1:
                for keyword in phone_keywords:
                    if keyword in header or keyword in header_lower:
                        column_mapping['phone'] = idx
                        break

            # 匹配备注列
            if column_mapping['note'] == -1:
                for keyword in note_keywords:
                    if keyword in header or keyword in header_lower:
                        column_mapping['note'] = idx
                        break

        # 如果没有找到姓名列，尝试使用默认位置
        if column_mapping['name'] == -1:
            # 检查是否是导出格式（编号,姓名,手机号,备注...）
            if len(headers) >= 7 and ('编号' in headers[0] or 'code' in headers[0].lower() or 'id' in headers[0].lower()):
                # 导出文件格式
                column_mapping['id'] = 0    # 编号在第1列
                column_mapping['name'] = 1  # 姓名在第2列
                column_mapping['phone'] = 2  # 手机号在第3列
                column_mapping['note'] = 3  # 备注在第4列
            elif len(headers) >= 3:
                # 导入模板格式（没有编号列）
                column_mapping['name'] = 0  # 姓名在第1列
                column_mapping['phone'] = 1  # 手机号在第2列
                column_mapping['note'] = 2  # 备注在第3列
            else:
                raise HTTPException(
                    status_code=400,
                    detail="无法识别文件格式，请确保包含姓名、手机号、备注列"
                )

        # 验证姓名列是否找到
        if column_mapping['name'] == -1:
            raise HTTPException(
                status_code=400,
                detail="未找到姓名列，请确保CSV文件包含'姓名'列"
            )

        # 手机号和备注列是可选的，如果没找到则使用默认值
        if column_mapping['phone'] == -1:
            column_mapping['phone'] = column_mapping['name'] + 1

        if column_mapping['note'] == -1:
            column_mapping['note'] = column_mapping['name'] + 2

        return column_mapping

    def batch_import_participants(
        self,
        activity_id: str,
        file: UploadFile,
        user_id: str
    ) -> ParticipantBatchImportResult:
        """批量导入参与者，支持Excel和CSV格式"""
        # 检查权限
        self._check_activity_permission(activity_id, user_id)

        filename = file.filename or ""
        is_csv = filename.lower().endswith('.csv')
        is_excel = filename.lower().endswith(('.xlsx', '.xls'))

        if not is_csv and not is_excel:
            raise HTTPException(
                status_code=400,
                detail="不支持的文件格式，请上传CSV或Excel文件"
            )

        try:
            contents = file.file.read()

            if is_csv:
                return self._import_from_csv(activity_id, contents)
            else:
                return self._import_from_excel(activity_id, contents)

        except HTTPException:
            raise
        except Exception as e:
            self.db.rollback()
            raise HTTPException(status_code=400, detail=f"文件处理错误: {str(e)}")

    def _import_from_csv(
        self,
        activity_id: str,
        contents: bytes
    ) -> ParticipantBatchImportResult:
        """从CSV文件导入参与者

        支持两种格式：
        1. 导入模板格式：姓名,手机号,备注
        2. 导出文件格式：编号,姓名,手机号,备注,是否入场,入场时间,创建时间
        """
        total = success = failed = 0
        errors = []

        try:
            # 尝试不同的编码
            text_content = None
            for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'gb18030']:
                try:
                    text_content = contents.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue

            if text_content is None:
                raise HTTPException(
                    status_code=400,
                    detail="无法识别文件编码，请使用UTF-8或GBK编码保存CSV文件"
                )

            # 解析CSV
            csv_file = io.StringIO(text_content)
            reader = csv.reader(csv_file)

            # 读取标题行以确定格式
            header_row = next(reader, None)
            if not header_row:
                raise HTTPException(status_code=400, detail="文件为空或格式错误")

            # 智能识别列映射
            column_mapping = self._detect_csv_format(header_row)

            # 处理每一行
            for idx, row in enumerate(reader, start=2):
                if not row or not any(row):
                    continue

                total += 1

                # 根据列映射提取数据
                try:
                    # 提取编号（如果有）
                    participant_id = None
                    if column_mapping['id'] != -1 and len(row) > column_mapping['id'] and row[column_mapping['id']]:
                        participant_id = row[column_mapping['id']].strip()

                    name = row[column_mapping['name']].strip() if len(
                        row) > column_mapping['name'] and row[column_mapping['name']] else ""
                    phone = row[column_mapping['phone']].strip() if len(
                        row) > column_mapping['phone'] and row[column_mapping['phone']] else None
                    note = row[column_mapping['note']].strip() if len(
                        row) > column_mapping['note'] and row[column_mapping['note']] else None
                except IndexError:
                    errors.append(f"第{idx}行：列数不足，请检查文件格式")
                    failed += 1
                    continue

                # 验证姓名
                if not name:
                    errors.append(f"第{idx}行：姓名不能为空")
                    failed += 1
                    continue

                # 检查重复
                existing = self.db.query(Participant).filter(
                    and_(
                        Participant.activity_id == activity_id,
                        Participant.name == name
                    )
                ).first()

                if existing:
                    errors.append(f"第{idx}行：参与者 {name} 已存在")
                    failed += 1
                    continue

                # 创建参与者
                try:
                    # 如果提供了编号，使用提供的编号，否则自动生成
                    if participant_id:
                        code = participant_id
                    else:
                        code = self._generate_participant_code(activity_id)

                    participant = Participant(
                        activity_id=activity_id,
                        code=code,
                        name=name,
                        phone=phone if phone else None,
                        note=note if note else None
                    )
                    self.db.add(participant)
                    success += 1
                except Exception as e:
                    errors.append(f"第{idx}行：{str(e)}")
                    failed += 1

            # 提交事务
            if success > 0:
                self.db.commit()

            return ParticipantBatchImportResult(
                total=total,
                success=success,
                failed=failed,
                errors=errors[:10]  # 只返回前10个错误
            )

        except Exception as e:
            self.db.rollback()
            raise HTTPException(status_code=400, detail=f"CSV文件处理错误: {str(e)}")

    def _import_from_excel(
        self,
        activity_id: str,
        contents: bytes
    ) -> ParticipantBatchImportResult:
        """从Excel文件导入参与者

        支持两种格式：
        1. 导入模板格式：姓名,手机号,备注
        2. 导出文件格式：编号,姓名,手机号,备注,是否入场,入场时间,创建时间
        """
        total = success = failed = 0
        errors = []

        try:
            workbook = load_workbook(io.BytesIO(contents))
            worksheet = workbook.active

            if worksheet is None:
                return ParticipantBatchImportResult(
                    total=0, success=0, failed=0, errors=["未找到工作表"]
                )

            # 读取标题行以确定格式
            header_row = None
            for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
                header_row = [
                    str(cell) if cell is not None else "" for cell in row]
                break

            if not header_row:
                return ParticipantBatchImportResult(
                    total=0, success=0, failed=0, errors=["未找到标题行"]
                )

            # 智能识别列映射
            column_mapping = self._detect_csv_format(header_row)

            # 跳过标题行，从第二行开始
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue

                total += 1

                # 根据列映射提取数据
                try:
                    # 提取编号（如果有）
                    participant_id = None
                    if column_mapping['id'] != -1 and len(row) > column_mapping['id'] and row[column_mapping['id']]:
                        participant_id = str(row[column_mapping['id']]).strip()

                    name = str(row[column_mapping['name']]).strip() if len(
                        row) > column_mapping['name'] and row[column_mapping['name']] else ""
                    phone = str(row[column_mapping['phone']]).strip() if len(
                        row) > column_mapping['phone'] and row[column_mapping['phone']] else None
                    note = str(row[column_mapping['note']]).strip() if len(
                        row) > column_mapping['note'] and row[column_mapping['note']] else None
                except IndexError:
                    errors.append(f"第{idx}行：列数不足，请检查文件格式")
                    failed += 1
                    continue

                # 验证姓名
                if not name:
                    errors.append(f"第{idx}行：姓名不能为空")
                    failed += 1
                    continue

                # 检查重复
                existing = self.db.query(Participant).filter(
                    and_(
                        Participant.activity_id == activity_id,
                        Participant.name == name
                    )
                ).first()

                if existing:
                    errors.append(f"第{idx}行：参与者 {name} 已存在")
                    failed += 1
                    continue

                # 创建参与者
                try:
                    # 如果提供了编号，使用提供的编号，否则自动生成
                    if participant_id:
                        code = participant_id
                    else:
                        code = self._generate_participant_code(activity_id)

                    participant = Participant(
                        activity_id=activity_id,
                        code=code,
                        name=name,
                        phone=phone if phone else None,
                        note=note if note else None
                    )
                    self.db.add(participant)
                    success += 1
                except Exception as e:
                    errors.append(f"第{idx}行：{str(e)}")
                    failed += 1

            # 提交事务
            if success > 0:
                self.db.commit()

            return ParticipantBatchImportResult(
                total=total,
                success=success,
                failed=failed,
                errors=errors[:10]  # 只返回前10个错误
            )

        except Exception as e:
            self.db.rollback()
            raise HTTPException(
                status_code=400, detail=f"Excel文件处理错误: {str(e)}")

    def export_participants(self, activity_id: str, user_id: str) -> bytes:
        """导出参与者数据为CSV"""
        # 检查权限
        activity = self._check_activity_permission(activity_id, user_id)

        # 获取参与者数据
        participants = self.db.query(Participant).filter(
            Participant.activity_id == activity_id
        ).order_by(Participant.code).all()

        # 创建CSV内容
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)

        # 写入标题行
        headers = ["编号", "姓名", "手机号", "备注", "是否入场", "入场时间", "创建时间"]
        writer.writerow(headers)

        # 写入数据行
        for participant in participants:
            # 处理手机号
            phone_value = getattr(participant, 'phone') or ""

            # 处理备注
            note_value = getattr(participant, 'note') or ""

            # 处理入场状态
            checked_in_value = getattr(participant, 'checked_in', False)
            checked_in_text = "是" if checked_in_value else "否"

            # 处理入场时间
            checked_in_at_value = getattr(participant, 'checked_in_at', None)
            if checked_in_at_value:
                checkin_time = checked_in_at_value.strftime(
                    "%Y-%m-%d %H:%M:%S")
            else:
                checkin_time = ""

            # 处理创建时间
            created_at_value = getattr(participant, 'created_at', None)
            if created_at_value:
                created_time = created_at_value.strftime("%Y-%m-%d %H:%M:%S")
            else:
                created_time = ""

            # 写入行数据
            writer.writerow([
                str(participant.code),
                str(participant.name),
                str(phone_value),
                str(note_value),
                checked_in_text,
                checkin_time,
                created_time
            ])

        # 转换为字节
        csv_content = output.getvalue()
        output.close()

        # 添加BOM以确保Excel正确显示中文
        return '\ufeff'.encode('utf-8') + csv_content.encode('utf-8')

    def generate_participant_link(self, participant_id: str, user_id: str) -> dict:
        """生成参与者链接参数

        返回参与者的活动ID和编号，用于前端构建链接
        """
        participant = self.db.query(Participant).filter(
            Participant.id == participant_id).first()
        if not participant:
            raise HTTPException(status_code=404, detail="参与者不存在")

        # 检查权限
        self._check_activity_permission(str(participant.activity_id), user_id)

        return {
            "activityId": str(participant.activity_id),
            "participantCode": str(participant.code)
        }

    def generate_participant_qrcode(self, participant_id: str, user_id: str) -> bytes:
        """生成参与者二维码（简化版本）"""
        # 简化实现，返回空字节
        return b"QR code generation not implemented yet"

    def participant_enter(
        self,
        activity_id: str,
        participant_code: str,
        device_fingerprint: Optional[str] = None
    ) -> tuple[dict, dict]:
        """参与者入场"""
        # 查找参与者
        participant = self.db.query(Participant).filter(
            and_(
                Participant.activity_id == activity_id,
                Participant.code == participant_code
            )
        ).first()

        if not participant:
            raise HTTPException(status_code=404, detail="参与者不存在或编号错误")

        # 更新入场状态（简化实现）
        # 注意：这里不直接修改SQLAlchemy对象的属性，而是使用update方法
        self.db.query(Participant).filter(Participant.id == participant.id).update({
            "checked_in": True,
            "checked_in_at": datetime.now(timezone.utc),
            "device_fingerprint": device_fingerprint
        })
        self.db.commit()

        # 获取活动信息
        activity = self.db.query(Activity).filter(
            Activity.id == activity_id).first()

        # 获取当前辩题
        current_debate = None
        current_debate_id = getattr(
            activity, 'current_debate_id', None) if activity else None
        if activity and current_debate_id:
            current_debate_obj = self.db.query(Debate).filter(
                Debate.id == current_debate_id
            ).first()
            if current_debate_obj:
                current_debate = {
                    "id": str(current_debate_obj.id),
                    "title": str(current_debate_obj.title),
                    "status": current_debate_obj.status.value
                }

        activity_info = {
            "activity": {
                "id": str(activity.id) if activity else "",
                "name": str(activity.name) if activity else "",
                "status": activity.status.value if activity else "unknown",
                "current_debate": current_debate
            },
            "participant": {
                "id": str(participant.id),
                "code": str(participant.code),
                "name": str(participant.name)
            }
        }

        # 检查当前辩题的投票状态
        has_voted = False
        vote_position = None
        voted_at = None
        remaining_changes = 3  # 默认值

        if activity and current_debate_id:
            # 查询当前辩题的投票记录
            current_vote = self.db.query(Vote).filter(
                and_(
                    Vote.participant_id == participant.id,
                    Vote.debate_id == current_debate_id
                )
            ).first()

            if current_vote:
                has_voted = True
                position_attr = getattr(current_vote, 'position', None)
                vote_position = position_attr.value if position_attr else None
                created_attr = getattr(current_vote, 'created_at', None)
                voted_at = created_attr.isoformat() if created_attr else None

            # 从活动设置中获取最大改票次数
            settings = getattr(activity, 'settings', None)
            if settings:
                remaining_changes = settings.get(
                    'max_vote_changes', settings.get('maxVoteChanges', 3))

                # 如果已投票，计算剩余改票次数
                if current_vote:
                    change_count = getattr(current_vote, 'change_count', 0)
                    remaining_changes = max(
                        0, remaining_changes - change_count)

        vote_status = {
            "has_voted": has_voted,
            "position": vote_position,
            "voted_at": voted_at,
            "remaining_changes": remaining_changes,
            "can_vote": True,
            "can_change": True
        }

        return activity_info, vote_status
