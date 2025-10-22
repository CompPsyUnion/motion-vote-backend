"""统计数据服务 - 基于Redis的实时统计 + 报表导出"""

import asyncio
import csv
import io
import json
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)
from sqlalchemy import and_, desc, func
from sqlalchemy.orm import Session

from src.core.redis import get_redis
from src.core.websocket_manager import broadcast_statistics_update
from src.models.activity import Activity, Collaborator
from src.models.debate import Debate
from src.models.user import User
from src.models.vote import Participant, Vote
from src.schemas.statistics import (ActivityReport, ActivitySummary,
                                    ActivityType, DashboardData, DebateResult,
                                    DebateStats, ExportType, RealTimeStats,
                                    RecentActivity, TimelinePoint, VoteStats)
from src.schemas.user import UserRole


class StatisticsService:
    """统计服务 - 集成Redis缓存和报表功能"""

    # 后台同步任务
    _sync_task: Optional[asyncio.Task] = None
    _last_broadcast: Dict[str, float] = {}  # activity_id -> timestamp

    def __init__(self, db: Session):
        self.db = db
        self.redis = get_redis()

        # 启动后台同步和广播任务
        if StatisticsService._sync_task is None:
            StatisticsService._sync_task = asyncio.create_task(
                self._background_worker()
            )

    # ============ Redis Key 生成 ============

    def _stats_key(self, activity_id: str) -> str:
        """活动统计数据的Redis key"""
        return f"stats:{activity_id}"

    def _debate_stats_key(self, debate_id: str) -> str:
        """辩题统计数据的Redis key"""
        return f"debate_stats:{debate_id}"

    def _dirty_activities_key(self) -> str:
        """需要同步到数据库的活动ID集合"""
        return "sync:dirty_activities"

    # ============ 权限检查 ============

    def _check_activity_permission(self, activity_id: str, user) -> Activity:
        """检查用户对活动的权限（使用 User 对象）

        Args:
            activity_id: 活动ID
            user: User 对象（包含 id 和 role）

        Returns:
            Activity: 活动对象
        """
        activity = self.db.query(Activity).filter(
            Activity.id == activity_id).first()
        if not activity:
            raise HTTPException(status_code=404, detail="Activity not found")

        # 支持传入 user 对象 或 user_id 字符串
        user_id = None
        user_obj = None
        from src.models.user import User as UserModel

        if isinstance(user, UserModel):
            user_obj = user
            user_id = str(user.id)
        else:
            user_id = str(user)

        # 如果传入了 user 对象并且是管理员，允许访问
        if user_obj is not None and str(user_obj.role) == "UserRole.admin":
            return activity

        # 检查是否为所有者
        if str(activity.owner_id) == user_id:
            return activity

        # 检查是否是协作者
        collaborator = self.db.query(Collaborator).filter(
            Collaborator.activity_id == activity_id,
            Collaborator.user_id == user_id
        ).first()

        if not collaborator:
            raise HTTPException(
                status_code=403,
                detail="You don't have permission to access this activity"
            )

        return activity

    # ============ Redis 缓存核心方法 ============

    async def get_activity_statistics(self, activity_id: str) -> Dict[str, Any]:
        """获取活动实时统计数据（优先从Redis读取）"""
        cache_key = self._stats_key(activity_id)

        # 尝试从Redis获取
        cached_data = self.redis.get(cache_key)  # type: ignore
        if cached_data and isinstance(cached_data, str):
            return json.loads(cached_data)

        # 缓存未命中，从数据库加载并缓存
        stats = await self._load_statistics_from_db(activity_id)
        self.redis.setex(  # type: ignore
            cache_key,
            300,  # 5分钟过期
            json.dumps(stats, ensure_ascii=False)
        )

        return stats

    async def get_debate_statistics(self, debate_id: str) -> Dict[str, Any]:
        """获取辩题实时统计数据（优先从Redis读取）"""
        cache_key = self._debate_stats_key(debate_id)

        # 尝试从Redis获取
        cached_data = self.redis.get(cache_key)  # type: ignore
        if cached_data and isinstance(cached_data, str):
            return json.loads(cached_data)

        # 缓存未命中，从数据库加载
        stats = await self._load_debate_stats_from_db(debate_id)
        self.redis.setex(  # type: ignore
            cache_key,
            300,  # 5分钟过期
            json.dumps(stats, ensure_ascii=False)
        )

        return stats

    async def update_statistics_cache(self, activity_id: str, debate_id: Optional[str] = None):
        """更新统计数据缓存（投票后调用）"""
        # 更新活动统计
        stats = await self._load_statistics_from_db(activity_id)
        cache_key = self._stats_key(activity_id)
        self.redis.setex(  # type: ignore
            cache_key,
            300,
            json.dumps(stats, ensure_ascii=False)
        )

        # 如果指定了辩题，也更新辩题统计
        if debate_id:
            debate_stats = await self._load_debate_stats_from_db(debate_id)
            debate_cache_key = self._debate_stats_key(debate_id)
            self.redis.setex(  # type: ignore
                debate_cache_key,
                300,
                json.dumps(debate_stats, ensure_ascii=False)
            )

        # 标记为需要广播
        await self._schedule_broadcast(activity_id, stats)

    async def _schedule_broadcast(self, activity_id: str, stats: Dict[str, Any]):
        """调度统计数据广播（防抖：1秒内最多广播一次）"""
        current_time = datetime.now(timezone.utc).timestamp()
        last_time = StatisticsService._last_broadcast.get(activity_id, 0)

        # 如果距离上次广播不到1秒，等待
        if current_time - last_time < 1.0:
            # 延迟广播
            await asyncio.sleep(1.0 - (current_time - last_time))

        # 广播统计数据
        await broadcast_statistics_update(activity_id, stats)
        StatisticsService._last_broadcast[activity_id] = datetime.now(
            timezone.utc).timestamp()

    async def _load_statistics_from_db(self, activity_id: str) -> Dict[str, Any]:
        """从数据库加载统计数据"""
        # 获取活动信息
        activity = self.db.query(Activity).filter(
            Activity.id == activity_id).first()
        if not activity:
            return {}

        # 获取当前辩题
        current_debate = None
        current_debate_stats = None

        if hasattr(activity, 'current_debate_id') and activity.current_debate_id is not None:
            debate = self.db.query(Debate).filter(
                Debate.id == activity.current_debate_id
            ).first()

            if debate:
                current_debate = {
                    "id": str(debate.id),
                    "title": str(debate.title),
                    "proDescription": str(debate.pro_description) if debate.pro_description is not None else None,
                    "conDescription": str(debate.con_description) if debate.con_description is not None else None,
                    "background": str(debate.background) if debate.background is not None else None,
                    "status": str(debate.status.value) if hasattr(debate.status, 'value') else str(debate.status),
                    "order": debate.order,
                    "activityId": str(debate.activity_id)
                }

                # 获取当前辩题的投票统计
                current_debate_stats = await self._get_debate_vote_stats_from_redis(str(debate.id))

        # 总参与人数
        total_participants = self.db.query(Participant).filter(
            Participant.activity_id == activity_id
        ).count()

        # 已入场人数
        checked_in_participants = self.db.query(Participant).filter(
            and_(
                Participant.activity_id == activity_id,
                Participant.checked_in == True
            )
        ).count()

        # 在线人数（简化版：等于已入场）
        online_participants = checked_in_participants

        # 总投票数
        total_votes = self.db.query(Vote).join(
            Participant, Vote.participant_id == Participant.id
        ).filter(Participant.activity_id == activity_id).count()

        # 投票率
        vote_rate = 0.0
        if checked_in_participants > 0 and current_debate_stats:
            current_votes = current_debate_stats.get('totalVotes', 0)
            vote_rate = round(current_votes / checked_in_participants * 100, 2)

        return {
            "activityId": str(activity.id),
            "activityName": str(activity.name),
            "activityStatus": str(activity.status.value) if hasattr(activity.status, 'value') else str(activity.status),
            "currentDebate": current_debate,
            "currentDebateStats": current_debate_stats,
            "realTimeStats": {
                "totalParticipants": total_participants,
                "checkedInParticipants": checked_in_participants,
                "onlineParticipants": online_participants,
                "totalVotes": total_votes,
                "voteRate": vote_rate
            },
            "timestamp": datetime.now(timezone.utc).isoformat()
        }

    async def _get_debate_vote_stats_from_redis(self, debate_id: str) -> Dict[str, Any]:
        """获取辩题的投票统计（从Redis或数据库）"""
        # 统计各立场的票数
        vote_counts = self.db.query(
            Vote.position,
            func.count(Vote.id).label('count')
        ).filter(Vote.debate_id == debate_id).group_by(Vote.position).all()

        pro_votes = 0
        con_votes = 0
        abstain_votes = 0

        for position, count in vote_counts:
            if position == 'pro':
                pro_votes = count
            elif position == 'con':
                con_votes = count
            elif position == 'abstain':
                abstain_votes = count

        total_votes = pro_votes + con_votes + abstain_votes

        # 计算百分比
        pro_percentage = round(pro_votes / total_votes *
                               100, 2) if total_votes > 0 else 0.0
        con_percentage = round(con_votes / total_votes *
                               100, 2) if total_votes > 0 else 0.0
        abstain_percentage = round(
            abstain_votes / total_votes * 100, 2) if total_votes > 0 else 0.0

        return {
            "debateId": debate_id,
            "proVotes": pro_votes,
            "conVotes": con_votes,
            "abstainVotes": abstain_votes,
            "totalVotes": total_votes,
            "proPercentage": pro_percentage,
            "conPercentage": con_percentage,
            "abstainPercentage": abstain_percentage
        }

    async def _load_debate_stats_from_db(self, debate_id: str) -> Dict[str, Any]:
        """从数据库加载辩题统计"""
        debate = self.db.query(Debate).filter(Debate.id == debate_id).first()
        if not debate:
            return {}

        vote_stats = await self._get_debate_vote_stats_from_redis(debate_id)

        return {
            "debateId": debate_id,
            "debateTitle": str(debate.title),
            "debateStatus": str(debate.status.value) if hasattr(debate.status, 'value') else str(debate.status),
            "voteStats": vote_stats,
            "timestamp": datetime.now(timezone.utc).isoformat()
        }

    async def _background_worker(self):
        """后台同步任务：每2.5秒刷新活跃活动的统计缓存"""
        while True:
            try:
                await asyncio.sleep(2.5)

                # 获取所有正在进行中的活动
                from src.schemas.activity import ActivityStatus
                from src.core.database import SessionLocal

                db = SessionLocal()

                try:
                    active_activities = db.query(Activity).filter(
                        Activity.status == ActivityStatus.ongoing
                    ).all()

                    for activity in active_activities:
                        try:
                            # 重新加载统计数据到缓存
                            stats = await self._load_statistics_from_db(str(activity.id))

                            # 更新Redis缓存（延长TTL）
                            cache_key = self._stats_key(str(activity.id))
                            self.redis.setex(  # type: ignore
                                cache_key,
                                300,  # 5分钟TTL
                                json.dumps(stats, ensure_ascii=False)
                            )

                        except Exception as e:
                            print(
                                f"Failed to refresh stats for activity {activity.id}: {e}")

                finally:
                    db.close()

            except Exception as e:
                print(f"Statistics cache sync error: {e}")
                await asyncio.sleep(5)

    # ============ Dashboard 数据（兼容旧接口）============

    def get_dashboard_data(self, activity_id: str, user: User) -> DashboardData:
        """获取实时数据看板"""
        # 检查权限
        activity = self._check_activity_permission(activity_id, user)

        # 获取实时统计数据
        real_time_stats = self._get_real_time_stats(activity_id)

        # 获取辩题统计
        debate_stats = self._get_debate_stats(activity_id)

        # 获取最近活动
        recent_activity = self._get_recent_activity(activity_id)

        # 获取当前辩题
        current_debate = None
        if hasattr(activity, 'current_debate_id') and activity.current_debate_id is not None:
            debate = self.db.query(Debate).filter(
                Debate.id == activity.current_debate_id).first()
            current_debate = str(debate.title) if debate else None

        return DashboardData(
            activityId=str(activity.id),
            activityName=str(activity.name),
            activityStatus=str(activity.status.value),
            currentDebate=current_debate,
            realTimeStats=real_time_stats,
            debateStats=debate_stats,
            recentActivity=recent_activity
        )

    def _get_real_time_stats(self, activity_id: str) -> RealTimeStats:
        """获取实时统计数据（优先从Redis缓存读取）"""
        # 尝试从Redis获取
        cache_key = self._stats_key(activity_id)
        cached_data = self.redis.get(cache_key)  # type: ignore

        if cached_data and isinstance(cached_data, str):
            stats = json.loads(cached_data)
            real_time_stats = stats.get('realTimeStats', {})
            return RealTimeStats(
                totalParticipants=real_time_stats.get('totalParticipants', 0),
                checkedInParticipants=real_time_stats.get(
                    'checkedInParticipants', 0),
                onlineParticipants=real_time_stats.get(
                    'onlineParticipants', 0),
                totalVotes=real_time_stats.get('totalVotes', 0),
                voteRate=real_time_stats.get('voteRate', 0.0)
            )

        # 缓存未命中，从数据库查询
        # 总参与人数
        total_participants = self.db.query(Participant).filter(
            Participant.activity_id == activity_id
        ).count()

        # 已入场人数
        checked_in_participants = self.db.query(Participant).filter(
            and_(
                Participant.activity_id == activity_id,
                Participant.checked_in == True
            )
        ).count()

        # 在线人数（简化：等于已入场人数）
        online_participants = checked_in_participants

        # 总投票数
        total_votes = self.db.query(Vote).join(
            Participant, Vote.participant_id == Participant.id
        ).filter(Participant.activity_id == activity_id).count()

        # 投票率
        vote_rate = 0.0
        if checked_in_participants > 0:
            # 获取当前辩题的投票数
            current_debate_votes = self.db.query(Vote).join(
                Participant, Vote.participant_id == Participant.id
            ).join(
                Debate, Vote.debate_id == Debate.id
            ).filter(
                and_(
                    Participant.activity_id == activity_id,
                    Debate.activity_id == activity_id,
                    Debate.status == 'ongoing'
                )
            ).count()

            vote_rate = round(current_debate_votes /
                              checked_in_participants * 100, 2)

        return RealTimeStats(
            totalParticipants=total_participants,
            checkedInParticipants=checked_in_participants,
            onlineParticipants=online_participants,
            totalVotes=total_votes,
            voteRate=vote_rate
        )

    def _get_debate_stats(self, activity_id: str) -> List[DebateStats]:
        """获取辩题统计"""
        debates = self.db.query(Debate).filter(
            Debate.activity_id == activity_id
        ).order_by(Debate.order).all()

        stats = []
        for debate in debates:
            vote_results = self._get_vote_results(str(debate.id))

            # 计算投票率
            checked_in_count = self.db.query(Participant).filter(
                and_(
                    Participant.activity_id == activity_id,
                    Participant.checked_in == True
                )
            ).count()

            vote_rate = 0.0
            if checked_in_count > 0:
                vote_rate = round(vote_results.total_votes /
                                  checked_in_count * 100, 2)

            stats.append(DebateStats(
                debateId=str(debate.id),
                debateTitle=str(debate.title),
                VoteStats=vote_results,
                voteRate=vote_rate
            ))

        return stats

    def _get_vote_results(self, debate_id: str) -> VoteStats:
        """获取投票结果"""
        from src.models.vote import VoteHistory

        # 统计各立场的当前票数
        vote_counts = self.db.query(
            Vote.position,
            func.count(Vote.id).label('count')
        ).filter(Vote.debate_id == debate_id).group_by(Vote.position).all()

        pro_votes = 0
        con_votes = 0
        abstain_votes = 0

        for position, count in vote_counts:
            if position == 'pro':
                pro_votes = count
            elif position == 'con':
                con_votes = count
            elif position == 'abstain':
                abstain_votes = count

        total_votes = pro_votes + con_votes + abstain_votes

        # 计算初始票数：从VoteHistory中获取每个投票的第一个位置
        votes_with_history = self.db.query(Vote).filter(
            Vote.debate_id == debate_id).all()

        pro_previous_votes = 0
        con_previous_votes = 0
        abstain_previous_votes = 0

        # 统计从各方跑票到其他方的人数
        # pro_to_con: 从正方跑到反方的人数
        # pro_to_abstain: 从正方跑到弃权的人数
        # con_to_pro: 从反方跑到正方的人数
        # con_to_abstain: 从反方跑到弃权的人数
        # abstain_to_pro: 从弃权跑到正方的人数
        # abstain_to_con: 从弃权跑到反方的人数
        pro_to_con = 0
        pro_to_abstain = 0
        con_to_pro = 0
        con_to_abstain = 0
        abstain_to_pro = 0
        abstain_to_con = 0

        for vote in votes_with_history:
            # 获取该投票的最早历史记录
            first_history = self.db.query(VoteHistory).filter(
                VoteHistory.vote_id == vote.id
            ).order_by(VoteHistory.created_at.asc()).first()

            if first_history:
                # 使用历史记录中的第一个位置
                initial_position = getattr(first_history, 'new_position')
            else:
                # 没有历史记录，说明从未改过票，使用当前位置
                initial_position = getattr(vote, 'position')

            # 当前位置
            current_position = getattr(vote, 'position')

            # 统计初始票数
            if initial_position == 'pro':
                pro_previous_votes += 1
            elif initial_position == 'con':
                con_previous_votes += 1
            elif initial_position == 'abstain':
                abstain_previous_votes += 1

            # 统计跑票情况（初始位置和当前位置不同）
            if initial_position != current_position:
                if initial_position == 'pro' and current_position == 'con':
                    pro_to_con += 1
                elif initial_position == 'pro' and current_position == 'abstain':
                    pro_to_abstain += 1
                elif initial_position == 'con' and current_position == 'pro':
                    con_to_pro += 1
                elif initial_position == 'con' and current_position == 'abstain':
                    con_to_abstain += 1
                elif initial_position == 'abstain' and current_position == 'pro':
                    abstain_to_pro += 1
                elif initial_position == 'abstain' and current_position == 'con':
                    abstain_to_con += 1

        # 计算弃权百分比
        abstain_percentage = (abstain_votes / total_votes *
                              100) if total_votes > 0 else 0

        # 计算得分（根据新规则）
        # 正方得分 = 反方到正方人数/反方初始人数 * 1000 + 中立到正方人数/中立初始人数 * 500
        pro_score = 0.0
        if con_previous_votes > 0:
            pro_score += (con_to_pro / con_previous_votes * 1000)
        if abstain_previous_votes > 0:
            pro_score += (abstain_to_pro / abstain_previous_votes * 500)

        # 反方得分 = 正方到反方人数/正方初始人数 * 1000 + 中立到反方人数/中立初始人数 * 500
        con_score = 0.0
        if pro_previous_votes > 0:
            con_score += (pro_to_con / pro_previous_votes * 1000)
        if abstain_previous_votes > 0:
            con_score += (abstain_to_con / abstain_previous_votes * 500)

        # 判定获胜方
        winner = None
        if total_votes > 0:
            if pro_score > con_score:
                winner = "pro"
            elif con_score > pro_score:
                winner = "con"
            else:
                winner = "tie"

        return VoteStats.model_validate({
            "debateId": debate_id,
            "proVotes": pro_votes,
            "conVotes": con_votes,
            "abstainVotes": abstain_votes,
            "totalVotes": total_votes,
            "proPreviousVotes": pro_previous_votes,
            "conPreviousVotes": con_previous_votes,
            "abstainPreviousVotes": abstain_previous_votes,
            "proToConVotes": pro_to_con,
            "conToProVotes": con_to_pro,
            "abstainToProVotes": abstain_to_pro,
            "abstainToConVotes": abstain_to_con,
            "abstainPercentage": round(abstain_percentage, 2),
            "proScore": round(pro_score, 2),
            "conScore": round(con_score, 2),
            "winner": winner,
            "isLocked": False,
            "lockedAt": None
        })

    def _get_recent_activity(self, activity_id: str, limit: int = 10) -> List[RecentActivity]:
        """获取最近活动"""
        # 这里是简化实现，实际应该从活动日志表获取
        activities = []

        # 获取最近的投票活动
        recent_votes = self.db.query(Vote).join(
            Participant, Vote.participant_id == Participant.id
        ).filter(Participant.activity_id == activity_id).order_by(
            desc(Vote.created_at)
        ).limit(limit).all()

        for vote in recent_votes:
            participant = self.db.query(Participant).filter(
                Participant.id == vote.participant_id
            ).first()
            debate = self.db.query(Debate).filter(
                Debate.id == vote.debate_id
            ).first()

            if participant and debate:
                activities.append(RecentActivity(
                    type=ActivityType.VOTE_CAST,
                    timestamp=getattr(vote, 'created_at', datetime.min) if getattr(
                        vote, 'created_at', None) is not None else datetime.min,
                    description=f"参与者 {participant.code} 对辩题「{debate.title}」投票：{vote.position}"
                ))

        return sorted(activities, key=lambda x: x.timestamp, reverse=True)[:limit]

    def get_activity_report(self, activity_id: str, user: User) -> ActivityReport:
        """获取活动报告"""
        # 检查权限
        activity = self._check_activity_permission(activity_id, user)

        # 获取活动摘要
        summary = self._get_activity_summary(activity_id)

        # 获取辩题结果
        debate_results = self._get_debate_results(activity_id)

        return ActivityReport(
            activityId=str(activity.id),
            activityName=str(activity.name),
            createdAt=getattr(activity, 'created_at', datetime.min) if getattr(
                activity, 'created_at', None) is not None else datetime.min,
            startedAt=getattr(activity, 'start_time', None),
            endedAt=getattr(activity, 'end_time', None),
            summary=summary,
            debateResults=debate_results,
            generatedAt=datetime.now(timezone.utc)
        )

    def _get_activity_summary(self, activity_id: str) -> ActivitySummary:
        """获取活动摘要"""
        # 总参与人数
        total_participants = self.db.query(Participant).filter(
            Participant.activity_id == activity_id
        ).count()

        # 总投票数
        total_votes = self.db.query(Vote).join(
            Participant, Vote.participant_id == Participant.id
        ).filter(Participant.activity_id == activity_id).count()

        # 辩题总数
        total_debates = self.db.query(Debate).filter(
            Debate.activity_id == activity_id
        ).count()

        # 平均投票率
        average_vote_rate = 0.0
        if total_participants > 0 and total_debates > 0:
            average_vote_rate = round(
                total_votes / (total_participants * total_debates) * 100, 2)

        # 活动时长
        activity = self.db.query(Activity).filter(
            Activity.id == activity_id).first()
        duration = 0
        if activity and hasattr(activity, 'start_time') and hasattr(activity, 'end_time'):
            if activity.start_time is not None and activity.end_time is not None:
                # 注意：这里使用计划时间，而不是实际执行时间
                duration = int(
                    (activity.end_time - activity.start_time).total_seconds() / 60)

        return ActivitySummary(
            totalParticipants=total_participants,
            totalVotes=total_votes,
            totalDebates=total_debates,
            averageVoteRate=average_vote_rate,
            duration=duration
        )

    def _get_debate_results(self, activity_id: str) -> List[DebateResult]:
        """获取辩题结果"""
        debates = self.db.query(Debate).filter(
            Debate.activity_id == activity_id
        ).order_by(Debate.order).all()

        results = []
        for debate in debates:
            vote_results = self._get_vote_results(str(debate.id))
            timeline = self._get_debate_timeline(str(debate.id))

            # 计算辩题持续时间
            duration = 0
            started_at = getattr(debate, 'started_at', None)
            ended_at = getattr(debate, 'ended_at', None)

            if started_at and ended_at:
                # 使用实际的开始和结束时间
                duration = int((ended_at - started_at).total_seconds() / 60)
            elif hasattr(debate, 'estimated_duration') and debate.estimated_duration is not None:
                # 如果没有实际时间，使用预估时长
                if hasattr(debate.estimated_duration, 'value'):
                    duration = debate.estimated_duration.value
                else:
                    duration = debate.estimated_duration

            results.append(DebateResult(
                debateId=str(debate.id),
                debateTitle=str(debate.title),
                debateOrder=int(getattr(debate, "order", 0)),
                results=vote_results,
                timeline=timeline,
                duration=int(duration) if isinstance(
                    duration, (int, float, str)) else 0
            ))

        return results

    def _get_debate_timeline(self, debate_id: str) -> List[TimelinePoint]:
        """获取辩题投票时间线"""
        # 简化实现：获取投票时间点
        votes = self.db.query(Vote).filter(
            Vote.debate_id == debate_id
        ).order_by(Vote.created_at).all()

        timeline = []
        pro_count = 0
        con_count = 0
        abstain_count = 0

        for vote in votes:
            if getattr(vote, 'position', None) == 'pro':
                pro_count += 1
            elif getattr(vote, 'position', None) == 'con':
                con_count += 1
            elif getattr(vote, 'position', None) == 'abstain':
                abstain_count += 1

            timeline.append(TimelinePoint(
                timestamp=getattr(vote, 'created_at', datetime.min) if getattr(
                    vote, 'created_at', None) is not None else datetime.min,
                proVotes=pro_count,
                conVotes=con_count,
                abstainVotes=abstain_count
            ))

        return timeline

    def export_data(self, activity_id: str, user: User, export_type: ExportType = ExportType.ALL) -> bytes:
        """导出原始数据为CSV格式"""
        # 检查权限
        self._check_activity_permission(activity_id, user)

        if export_type == ExportType.VOTES or export_type == ExportType.ALL:
            return self._export_votes_csv(activity_id)
        elif export_type == ExportType.CHANGES:
            return self._export_changes_csv(activity_id)
        elif export_type == ExportType.TIMELINE:
            return self._export_timeline_csv(activity_id)
        else:
            return self._export_all_csv(activity_id)

    def _export_votes_csv(self, activity_id: str) -> bytes:
        """导出投票数据为CSV"""
        # 获取投票数据
        votes_query = self.db.query(
            Participant.code.label('participant_code'),
            Participant.name.label('participant_name'),
            Debate.title.label('debate_title'),
            Debate.order.label('debate_order'),
            Vote.position,
            Vote.created_at.label('vote_time'),
            Vote.updated_at.label('last_updated')
        ).join(
            Participant, Vote.participant_id == Participant.id
        ).join(
            Debate, Vote.debate_id == Debate.id
        ).filter(
            Participant.activity_id == activity_id
        ).order_by(Vote.created_at)

        # 创建CSV
        output = io.StringIO()
        writer = csv.writer(output)

        # 写入表头
        writer.writerow([
            '参与者编号', '参与者姓名', '辩题标题', '辩题顺序',
            '投票立场', '投票时间', '最后更新时间'
        ])

        # 写入数据
        for vote in votes_query.all():
            position_map = {
                'pro': '正方',
                'con': '反方',
                'abstain': '弃权'
            }

            writer.writerow([
                vote.participant_code,
                vote.participant_name,
                vote.debate_title,
                vote.debate_order,
                position_map.get(vote.position, vote.position),
                vote.vote_time.strftime(
                    '%Y-%m-%d %H:%M:%S') if vote.vote_time else '',
                vote.last_updated.strftime(
                    '%Y-%m-%d %H:%M:%S') if vote.last_updated else ''
            ])

        # 添加BOM以确保Excel正确显示中文
        csv_content = output.getvalue()
        output.close()

        return '\ufeff'.encode('utf-8') + csv_content.encode('utf-8')

    def _export_changes_csv(self, activity_id: str) -> bytes:
        """导出投票变更记录为CSV（简化实现）"""
        # 这里应该从投票变更日志表获取数据，目前简化处理
        return self._export_votes_csv(activity_id)

    def _export_timeline_csv(self, activity_id: str) -> bytes:
        """导出时间线数据为CSV"""
        # 获取所有辩题的时间线数据
        debates = self.db.query(Debate).filter(
            Debate.activity_id == activity_id
        ).order_by(Debate.order).all()

        output = io.StringIO()
        writer = csv.writer(output)

        # 写入表头
        writer.writerow([
            '辩题标题', '辩题顺序', '时间点', '正方票数', '反方票数', '弃权票数', '总票数'
        ])

        for debate in debates:
            timeline = self._get_debate_timeline(str(debate.id))
            for point in timeline:
                total = point.pro_votes + point.con_votes + point.abstain_votes
                writer.writerow([
                    debate.title,
                    debate.order,
                    point.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    point.pro_votes,
                    point.con_votes,
                    point.abstain_votes,
                    total
                ])

        csv_content = output.getvalue()
        output.close()

        return '\ufeff'.encode('utf-8') + csv_content.encode('utf-8')

    def _export_all_csv(self, activity_id: str) -> bytes:
        """导出所有数据为CSV"""
        # 简化实现：返回投票数据
        return self._export_votes_csv(activity_id)

    def generate_pdf_report(self, activity_id: str, user: User) -> bytes:
        """生成PDF格式的活动报告"""
        # 获取报告数据
        report = self.get_activity_report(activity_id, user)

        # 创建PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()

        # 标题
        title = Paragraph(
            f"<b>{report.activity_name}</b><br/>活动报告", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 0.3*inch))

        # 活动摘要
        summary_data = [
            ['指标', '数值'],
            ['总参与人数', str(report.summary.total_participants)],
            ['总投票数', str(report.summary.total_votes)],
            ['辩题总数', str(report.summary.total_debates)],
            ['平均投票率', f"{report.summary.average_vote_rate}%"],
            ['活动时长(分钟)', str(report.summary.duration)],
        ]

        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))

        # 辩题结果
        story.append(Paragraph("<b>辩题投票结果</b>", styles['Heading2']))
        story.append(Spacer(1, 0.2*inch))

        for debate in report.debate_results:
            # 从 results 字段中获取投票数据
            pro_votes = debate.results.pro_votes
            con_votes = debate.results.con_votes
            abstain_votes = debate.results.abstain_votes
            total_votes = pro_votes + con_votes + abstain_votes
            pro_percentage = (pro_votes / total_votes *
                              100) if total_votes > 0 else 0

            debate_data = [
                ['辩题', debate.debate_title],
                ['赞成票', str(pro_votes)],
                ['反对票', str(con_votes)],
                ['弃权票', str(abstain_votes)],
                ['总票数', str(total_votes)],
                ['赞成率', f"{pro_percentage:.2f}%"],
            ]

            debate_table = Table(debate_data, colWidths=[2*inch, 3*inch])
            debate_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            story.append(debate_table)
            story.append(Spacer(1, 0.2*inch))

        # 生成时间
        story.append(Spacer(1, 0.5*inch))
        gen_time = f"生成时间: {report.generated_at.strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(gen_time, styles['Normal']))

        # 构建PDF
        doc.build(story)

        pdf_content = buffer.getvalue()
        buffer.close()

        return pdf_content

    def generate_excel_report(self, activity_id: str, user: User) -> bytes:
        """生成Excel格式的活动报告"""
        # 获取报告数据
        report = self.get_activity_report(activity_id, user)

        # 创建工作簿
        wb = Workbook()

        # 活动摘要工作表
        ws_summary = wb.active
        if ws_summary:
            ws_summary.title = "活动摘要"

            # 标题
            ws_summary['A1'] = report.activity_name
            ws_summary['A1'].font = Font(bold=True, size=16)
            ws_summary.merge_cells('A1:B1')

            # 摘要数据
            summary_headers = ['指标', '数值']
            ws_summary.append([])
            ws_summary.append(summary_headers)

            # 设置表头样式
            for col in ['A', 'B']:
                cell = ws_summary[f'{col}3']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            # 添加数据
            ws_summary.append(['总参与人数', report.summary.total_participants])
            ws_summary.append(['总投票数', report.summary.total_votes])
            ws_summary.append(['辩题总数', report.summary.total_debates])
            ws_summary.append(
                ['平均投票率', f"{report.summary.average_vote_rate}%"])
            ws_summary.append(['活动时长(分钟)', report.summary.duration])

            # 设置列宽
            ws_summary.column_dimensions['A'].width = 20
            ws_summary.column_dimensions['B'].width = 15

        # 辩题结果工作表
        ws_debates = wb.create_sheet(title="辩题结果")

        # 表头
        debate_headers = ['辩题标题', '顺序', '赞成票',
                          '反对票', '弃权票', '总票数', '赞成率', '投票率']
        ws_debates.append(debate_headers)

        # 设置表头样式
        for col_idx, header in enumerate(debate_headers, start=1):
            cell = ws_debates.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')

        # 添加辩题数据
        for debate in report.debate_results:
            # 从 results 字段中获取投票数据
            pro_votes = debate.results.pro_votes
            con_votes = debate.results.con_votes
            abstain_votes = debate.results.abstain_votes
            total_votes = pro_votes + con_votes + abstain_votes
            pro_percentage = (pro_votes / total_votes *
                              100) if total_votes > 0 else 0
            vote_rate = (total_votes / report.summary.total_participants *
                         100) if report.summary.total_participants > 0 else 0

            ws_debates.append([
                debate.debate_title,
                debate.debate_order,
                pro_votes,
                con_votes,
                abstain_votes,
                total_votes,
                f"{pro_percentage:.2f}%",
                f"{vote_rate:.2f}%"
            ])

        # 自动调整列宽
        for column in ws_debates.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_debates.column_dimensions[column_letter].width = adjusted_width

        # 生成时间
        if ws_summary:
            ws_summary.append([])
            ws_summary.append(
                ['生成时间', report.generated_at.strftime('%Y-%m-%d %H:%M:%S')])

        # 保存到字节流
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_content = buffer.getvalue()
        buffer.close()

        return excel_content


# 辅助函数：用于大屏实时数据
async def get_activity_statistics(db: Session, activity_id: str) -> dict:
    """
    获取活动的实时统计数据（用于大屏显示）
    Returns complete statistics for screen display
    """
    try:
        # 获取活动信息
        activity = db.query(Activity).filter(
            Activity.id == activity_id).first()
        if not activity:
            raise ValueError(f"Activity {activity_id} not found")

        # 获取当前辩题
        current_debate = None
        current_debate_data = None
        if hasattr(activity, 'current_debate_id') and activity.current_debate_id is not None:
            debate = db.query(Debate).filter(
                Debate.id == activity.current_debate_id
            ).first()
            if debate:
                current_debate = str(debate.id)
                current_debate_data = {
                    "id": str(debate.id),
                    "title": debate.title,
                    "description": debate.description,
                    "status": debate.status.value if hasattr(debate, 'status') else 'unknown'
                }

        # 获取所有辩题的投票统计
        debates = db.query(Debate).filter(
            Debate.activity_id == activity_id
        ).all()

        debate_statistics = []
        for debate in debates:
            # 正方票数
            affirmative_votes = db.query(Vote).filter(
                and_(
                    Vote.debate_id == debate.id,
                    Vote.side == 'affirmative'
                )
            ).count()

            # 反方票数
            negative_votes = db.query(Vote).filter(
                and_(
                    Vote.debate_id == debate.id,
                    Vote.side == 'negative'
                )
            ).count()

            # 总票数
            total_votes = affirmative_votes + negative_votes

            # 计算百分比
            affirmative_percentage = (
                affirmative_votes / total_votes * 100) if total_votes > 0 else 0
            negative_percentage = (
                negative_votes / total_votes * 100) if total_votes > 0 else 0

            debate_statistics.append({
                "debate_id": str(debate.id),
                "title": debate.title,
                "status": debate.status.value if hasattr(debate, 'status') else 'unknown',
                "affirmative_votes": affirmative_votes,
                "negative_votes": negative_votes,
                "total_votes": total_votes,
                "affirmative_percentage": round(affirmative_percentage, 2),
                "negative_percentage": round(negative_percentage, 2),
                "is_current": str(debate.id) == current_debate
            })

        # 参与者统计
        total_participants = db.query(Participant).filter(
            Participant.activity_id == activity_id
        ).count()

        checked_in_participants = db.query(Participant).filter(
            and_(
                Participant.activity_id == activity_id,
                Participant.checked_in == True
            )
        ).count()

        # 总投票数
        total_votes_count = db.query(Vote).join(Debate).filter(
            Debate.activity_id == activity_id
        ).count()

        return {
            "activity_id": str(activity_id),
            "activity_name": activity.name,
            "activity_status": activity.status.value if hasattr(activity, 'status') else 'unknown',
            "current_debate": current_debate_data,
            "participants": {
                "total": total_participants,
                "checked_in": checked_in_participants,
                "participation_rate": round(
                    (checked_in_participants / total_participants *
                     100) if total_participants > 0 else 0,
                    2
                )
            },
            "total_votes": total_votes_count,
            "debates": debate_statistics,
            "timestamp": datetime.now(timezone.utc).isoformat()
        }

    except Exception as e:
        print(f"Error getting activity statistics: {e}")
        raise


# ============ 全局服务实例 ============

_statistics_service: Optional[StatisticsService] = None


def get_statistics_service(db: Session) -> StatisticsService:
    """获取统计服务实例（统一入口）"""
    global _statistics_service
    if _statistics_service is None:
        _statistics_service = StatisticsService(db)
    return _statistics_service
